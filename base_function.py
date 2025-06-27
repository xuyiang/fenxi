
from urllib.parse import urlparse, parse_qs, unquote, urlunparse
from datetime import  datetime, timedelta, date
from time import sleep
import time


import re
import os

import pandas as pd
import numpy
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, load_workbook

from selenium.common import TimeoutException
from selenium import webdriver
from selenium.webdriver.support.ui import  WebDriverWait
from selenium.webdriver.support import  expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options as chrome_Options
from selenium.webdriver.chrome.service import Service  as chrome_Service
from selenium.webdriver.edge.options import  Options as edge_Options
from selenium.webdriver.edge.service import  Service as edge_Service


from openai import OpenAI
import json


import config as cfg
from config import  global_query_types, global_query_Chinese_name ,global_release_date_types
from config import  ms_edge_web_browser_driver, chrome_web_driver


########################################################################################################################
#创建驱动器
def initial_web_driver(browse_type):
    """
    返回指定的浏览器驱动器
    :param browse_type:"chrome"   google 浏览器，   “edge”  微软edge 浏览器
    :return:
    """
    if browse_type.lower() == "chrome":
        print("调用 google Chrome 浏览器")
        driver_options = chrome_Options()
        # 添加参数禁用保存密码提示
        driver_options.add_experimental_option("prefs", {
            "credentials_enable_service": False,
            "profile": {
                "password_manager_enabled": False
            }
        })


        service = chrome_Service(executable_path= chrome_web_driver)
    else:
        print("调用 Microsoft Edge 浏览器")
        driver_options = edge_Options()

        # 添加参数禁用保存密码提示
        driver_options.add_argument('--inprivate')

        service = edge_Service(executable_path= ms_edge_web_browser_driver)

    # 关闭自动化检测
    driver_options.add_experimental_option('excludeSwitches', ['enable-automation'])
    if browse_type.lower() == "chrome":
        driver = webdriver.Chrome(options=driver_options, service=service)
    else:
        driver = webdriver.Edge(options=driver_options, service=service)
    return driver


########################################################################################################################
#视窗显示设置
def set_document_zoom_scale(main_web_driver: webdriver, according :int, specify_scale=None):
    """
    设置页面的显示比例。
    :param main_web_driver:
    :param according: 1：横向最佳比例， 2:纵向比例  3：宽纵中的最小值。 其他值不做操作。
    :param specify_scale:此值不为空时，按照指定值设置页面显示比例. 如 0.8 ， 表示80%。
    :return:
    """
    if specify_scale is not None:
        specify_scale = int(specify_scale * 100)
        main_web_driver.execute_script(f"document.body.style.zoom='{specify_scale}%'")
        return
    window_size = main_web_driver.get_window_size()
    content_width = main_web_driver.execute_script("return document.body.scrollWidth;")
    width_scale = int(window_size['width']*1.0/content_width*100)
    content_height = main_web_driver.execute_script("return document.body.scrollHeight;")
    height_scale = int(window_size['height']*1.0/content_height*100)
    result_scale = 100
    match according:
        case 1:
            result_scale = width_scale
        case 2 :
            result_scale = height_scale
        case 3:
            result_scale = width_scale if width_scale < height_scale else height_scale
        case _:
            result_scale = 100
    result_scale = result_scale if result_scale <=100 else 100
    main_web_driver.execute_script(f"document.body.style.zoom='{result_scale}%'")





########################################################################################################################
#审核

#审核date ,datetime是否在范围内
def audit_date_time_range(audited_date, date_begin, date_end) -> bool:
    """
    判断日期范围是否在日期范围之内 。
    对于datetime,  date_begin<audited_date<=date_end。
    对于 date ， date_begin<=audited_date<=date_end。
    其他数据类型，返回False
    :param audited_date: 待比较的日期或时间
    :param date_begin:日期或时间起点
    :param date_end:日期或时间终点
    :return: 符合要求，返回 True, 否则，返回 False
    """
    if isinstance(audited_date,date):
        result = date_begin <= audited_date <= date_end
    elif isinstance(audited_date,datetime):
        result = date_begin < audited_date <= date_end
    else:
        result = False
    return result

#判断公司名称是否在公司列表名中
def is_available_company_name(available_company_2Darray,audited_company_name )->int:
    """
    #判断公司名称是否在公司列表名中
    :param available_company_2Darray: 单位名称列表， 包括公司名， 别名（基本不用）， 名字中排除的字， 销售单位， 客户代表["公司名称","别名","排除词组","销售部门","客户代表"]]
    :param audited_company_name:待审核公司名
    :return:存在， 返回符合公司审核名称的 在 available_company_2Darray的 行号。<0 为无效公司名称
    """
    if available_company_2Darray is None or len(available_company_2Darray)<1 :
        return -1
    for index in range(0, len(available_company_2Darray)):
        one_row = available_company_2Darray[index]
        if audit_company_name(audited_company_name,one_row[0],one_row[1], one_row[2]):
            return index
    return -1


#公司名称审核。 公司名称 in （company_name , alias_names）  , 并且不出现再 excluded_statements 中
def audit_company_name(audited_company_name, company_name, alias_names, excluded_statements) -> bool:
    """
    审核待核查公司名称是否符合名称审核要求。
    :param audited_company_name: 被审核公司名称
    :param company_name:  可以使用的公司名称
    :param alias_names:  可以使用的公司别名 ，必须是字符串类型,使用[;,，；] 分割    (*在实际工作中，为避免重复录入数据，或者 少录入数据，alias保持为空字符)
    :param excluded_statements:    公司名称中不能包含的词组，必须是字符串类型，使用[;,，；] 分割
    :return:  是否符合审核要求
    """
    if type(alias_names) != str:
        alias_names = f"{alias_names}"
    if type(excluded_statements) != str:
        excluded_statements = f"{excluded_statements}"
    pass_company_name_audit = True
    if audited_company_name.find(company_name.strip())  == -1:
        pass_company_name_audit = False
        #没有找到被审核公司名称,查看是否符合别名要求
        for one_alias_name in re.split(r'[;,，；]', alias_names):
            one_alias = one_alias_name.strip()
            if len(one_alias) == 0:  #空字符串傲寒在任意字符串中，必须排除
                continue
            if audited_company_name.find(one_alias.strip()) >-1:
                pass_company_name_audit = True
                break
    if not pass_company_name_audit:
        return False
    for one_excluded_statement in re.split(r'[;,，；]', excluded_statements) :
        one_exclude = one_excluded_statement.strip()
        if len(one_exclude) == 0: #空字符串傲寒在任意字符串中，必须排除
            continue
        if audited_company_name.find(one_exclude.strip())>-1:
            pass_company_name_audit = False
            break
    return pass_company_name_audit

#是否是有效的内容。用户可以自定义判断正则规则，re_express_list必须使用字符串列表形式。
def is_valid_content(content, min_len= 20, re_express_list=None)->bool:
    """
    判断content内容是否可用于人工智能分析。
    :param content:待判断的字符串
    :param min_len:  字符串最小长谷
    :param re_express_list: 规则正则表达式字符串 list。若为 None，使用预置的表达式。预置条件适用于剑鱼
    :return: 符合要求，返回True, 否则 False
    """
    if len(content)< min_len :
        return False

    if re_express_list is None:
        details_filter_re_pattern_expresses = ["^页面异常，未获得访问权限！", "^权限不足，无法查看", "^人工查询内容"]
    else:
        details_filter_re_pattern_expresses = re_express_list.copy()

    details_filter_re_pattern_list = []
    for one_re_pattern_express in details_filter_re_pattern_expresses:
        details_filter_re_pattern_list.append(re.compile(one_re_pattern_express,re.MULTILINE))
    re_match = False
    for one_re_patter in details_filter_re_pattern_list:
        re_match = one_re_patter.search(string=content)
        if re_match:  # 遍历所有的表达式
            break

    return not re_match


#判断dataframe 中是否存在指定的记录：
def is_record_exist(df_records, bid_type, title, tenderer, release_date, budget, web_hyperlink)->bool:
    """
    根据输入参数，判断是否存在相同记录
    :param df_records:
    :param bid_type:
    :param title:
    :param tenderer:
    :param release_date:
    :param budget: 预算如果为空， 则不比较预算金额
    :param web_hyperlink:
    :return:
    """
    if df_records is None :
        return False
    df_filter = df_records[df_records["采购单位"] == tenderer]
    if len(df_filter) == 0:
        return False

    str_query_1 = f'类型 ==@bid_type and 发布日期==@release_date and 标题 ==@title'
    df_filter = df_filter.query(str_query_1)
    if len(df_filter)==0:
        return False

    if budget is  not None:
        df_filter = df_filter[df_filter["预算金额(万)"]==budget]
        # 允许一定误差范围内视为相等（例如误差小于 1e-9）
        #df_filter = df_filter[numpy.isclose(df_filter['预算金额(万)'], budget, atol=1e-9)]
        if len(df_filter)==0:
            return False

    df_filter = df_filter[df_filter["网站链接"] == web_hyperlink]
    if len(df_filter)==0:
        return False
    #print(df_filter2.to_string(max_rows=None, max_cols=None))
    return True

########################################################################################################################
#格式转换
#将字符串转换为datetime， 不正确的字符串返回1900-01-01 00:00:00
def string_to_datetime(string_datetime) ->datetime:
    """
    把字符串转换为日期格式
    :param string_datetime:
    :return:
    """
    string_datetime = string_datetime.strip()  # 注意， 字符串当中的空格不能删除， 正确格式中有空格。 2025-01-01 00:02:08
    # 针对不带年的情况，添加当年的年份
    datetime_pattern = r'(^\d{2}-\d{2})'
    match = re.match(datetime_pattern,string_datetime)

    #预设一个异常或不正确情况下会返回的时间类型
    datetime_object = datetime.strptime("1900-01-01", "%Y-%m-%d")

    if match:
        string_datetime = f"{datetime.now().year}-{string_datetime}"
    #解析类型 2025-01-01 22:30:30的字符串
    datetime_pattern = r'(^\d{4}-\d{2}-\d{2}( \d{2}:\d{2}(:\d{2})?)?)'
    format_datetime = ["%Y-%m-%d","%Y-%m-%d %H:%M","%Y-%m-%d %H:%M:%S"]
    match = re.match(datetime_pattern,string_datetime)
    if match :
        counter = string_datetime.count(':')
        datetime_locator = counter
        try :
            datetime_object = datetime.strptime(match.group(1), format_datetime[datetime_locator])
        except Exception as e:
            print(f"解析时间出错！字符串：{string_datetime}")
    else:
        #不匹配，查看是否带小时或者分钟 ,此刻要删除其中的空格
        string_datetime = string_datetime.replace(" ","")
        delta_datetime_pattern = (r'^'
                                  r'(?:(?P<days>\d+)天)?'       # 天（可选）
                                  r'(?:(?P<hours>\d+)小时)?'  # 小时（可选）
                                  r'(?:(?P<minutes>\d+)分(钟)?)?'  # 分钟
                                  r'(?:(?P<seconds>\d+)\秒(钟)?)?'  # 秒（可选）
                                  r'(?P<direction>[前后])'  # 方向（必选）
                                  r'$')
        match = re.match(delta_datetime_pattern,string_datetime)
        if match:
            print(list(match.groupdict().items())[:-1])
            data = {k: int(v) if v else 0 for k, v in list(match.groupdict().items())[:-1]}
            if  match.group("direction") =="后":
                direction = -1
            else:
                direction = 1
            days = data["days"] * direction
            hours = data["hours"] * direction
            minutes = data["minutes"] * direction
            seconds = data["seconds"] * direction
            if days + hours + minutes + seconds == 0: #至少有一个数字
                match = False
            else:
                offset_day_hour_minute_second = timedelta(days= days, hours=hours, minutes=minutes , seconds= seconds)
                datetime_object = datetime.now().replace(microsecond=0) - offset_day_hour_minute_second
    if  not match:
        print(f"解析时间出错！字符串：{string_datetime}")
    return datetime_object

#将包含金额的字符串转换为以万为单位的数字。不符合要求的字符串返回 None
def calculate_amount(text:str):
    """
    计算金额，返回以万元为单位的数字。若字符串不满足要求，返回 None
    :param text: 形如： 1.4亿元,  返回 14000 1500万， 返回 1500， 2000元， 返回0.2
    :return:返回以万元为单位的数字。若字符串不满足要求，返回 None
    """
    text = text.replace(" ","")
    pattern = r'^(\d+\.?\d*)(?:(万|亿)(?:元)?|(元))$'
    match = re.match(pattern, text)
    if match:
        num = float(match.group(1))
        unit = match.group(2) or match.group(3)
    else:
        return None
    if unit =="亿":
        num *= 10000.0
    elif unit =="元":
        num /= 10000.0
    else:
        pass
    return num

#将URL中的参数对应的内容进行解码， 解决URL链接过长问题
def decode_keyword_in_url(url, keyword , is_ignore_case = True):
    """
    对URL关键词对应的值进行解码，解决链接过长的问题
    :param url:  需要被解码的URL
    :param keyword:  对应的参数
    :param is_ignore_case:  是否忽略大小写
    :return:  解码后的URL
    """
    # 解析URL
    parsed_url = urlparse(url)
    # 解析查询参数
    query_params = parse_qs(parsed_url.query, keep_blank_values=True)

    # 查找keyword参数（忽略大小写）
    keyword_key = None
    if is_ignore_case:
        keyword_value = keyword.lower()
    else:
        keyword_value = keyword

    for key in query_params:
        if is_ignore_case:
            key_value= key.lower()
        else:
            key_value =key

        if key_value == keyword_value:
            keyword_key = key
            break

    # 如果找到keyword参数
    if keyword_key:
        # 解码keyword参数
        decoded_keyword = unquote(query_params[keyword_key][0])

        # 更新查询参数
        query_params[keyword_key] = [decoded_keyword]

        # 重新构建查询字符串
        new_query = '&'.join([f"{k}={v[0]}" for k, v in query_params.items()])

        # 重新构建URL
        new_url = urlunparse((
            parsed_url.scheme,
            parsed_url.netloc,
            parsed_url.path,
            parsed_url.params,
            new_query,
            parsed_url.fragment
        ))

        return new_url
    else:
        # 如果没有keyword参数，返回原URL
        return url

#将字符串分解为两个日期字符串
def  resolve_date_range(date_range_string,re_pattern_string = r'(\d{4}/\d{2}/\d{2})-(\d{4}/\d{2}/\d{2})'):
    """
    匹配成功，返回结果组，否则返回None
    :param date_range_string: 待分解的字符串
    :param re_pattern_string:正则表达式。
    :return:
    """
    date_pattern = re.compile(re_pattern_string)
    match = date_pattern.match(date_range_string)
    if match:
        return match.groups()
    else :
        return None

#将dataframe指定列转换为浮点数列
def transform_column_to_float(df_dataframe,column_name):
    """
    将dataframe指定列转换为浮点数列
    :param df_dataframe:
    :param column_name: 需要转换的列
    :return:
    """
    if df_dataframe is None:
        return
    df_dataframe[column_name] = (
        df_dataframe[column_name]
        .str.replace(r"[^\d.]", "", regex=True)  # 清理字符
        .replace(r"^$", "0", regex=True)  # 空字符串替换为 0（可选）
        .pipe(pd.to_numeric, errors="coerce")  # 安全转换为 float，无效值变 NaN
    )

#使用正则表达式，对指定的列进行信息提取，变更
def extract_column_by_re(df_dataframe,  column_name, re_express,extract_index=0):
    """
    使用正则表达式，对指定的列进行信息提取，变更
    :param df_dataframe:
    :param column_name: 需要提取信息的列名
    :param re_express: 正则表达式。 提取URL示例：r'"(http(s)?://(.*?))"' ,，需要提取的部分用括号
    :param extract_index:  提取第几个适配组。 起始需要为0
    :return: 提取后的表达式
    """
    if df_dataframe is None:
        return
    re_pattern = re.compile(re_express)
    df_dataframe[column_name] = df_dataframe[column_name].str.extract(re_pattern)[extract_index].fillna('')

########################################################################################################################
#页面交互操作
#统一操作等待时间
def wait_after_operation(seconds):
    """
    统一各类暂停后的擦操作模式
    :param seconds:
    :return:
    """
    if seconds <0.001 :
        return
    sleep(seconds)

#让界面随着处理的数据进行滚动
def scroll_window_synchronize(browser_driver, current_record_num, total_record_num):
    """
    随着当前解析的记录的位置，同步滚动窗口
    :param browser_driver:
    :param current_record_num: 当前需要滚动到的记录序号
    :param total_record_num: 本页面记录总数
    :return:
    """
    if total_record_num <= 3:
        return
    if current_record_num <= 1:
        return
    if current_record_num == total_record_num:
        browser_driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        return
    page_height = browser_driver.execute_script("return document.body.scrollHeight;")
    current_scroll_x = browser_driver.execute_script("return window.pageXOffset;")
    current_scroll_y = browser_driver.execute_script("return window.pageYOffset;")
    new_scroll_y= int(current_scroll_y + (page_height-current_scroll_y) / (total_record_num - current_record_num + 2))
    browser_driver.execute_script(f"window.scrollTo({current_scroll_x}, {new_scroll_y});")

#点击超链接element , 如不能点击，尝试移动鼠标后再点击
def click_or_move_mouse(browse_driver, element_to_be_click, name_of_element="待点击控件",
                        time_before_move_mouse_and_smooth_scroll = cfg.time_before_move_mouse_and_smooth_scroll,
                        time_after_move_mouse_and_smooth_scroll = cfg.time_after_move_mouse_and_smooth_scroll)->bool:
    """
    点击，若失败，移动鼠标至控件，再次点击
    :param browse_driver:
    :param element_to_be_click:  将被点击的控件
    :param name_of_element:  将被点击的控件名，用户打印输出时，显示控件名称
    :param time_before_move_mouse_and_smooth_scroll: 移动鼠标前等候时间
    :param time_after_move_mouse_and_smooth_scroll: 移动鼠标至控件后，点击控件前的等待时间
    :return:
    """
    if element_to_be_click is None :
        return False
    try:
        element_to_be_click.click()
        print(f"点击{name_of_element}成功")
        return True
    except Exception as e:
        message = f'点击{name_of_element} 报错！ '
        to_qw_error(message, exception_info=e)
    print(f"尝试移动鼠标后再次点击 {name_of_element} ")
    wait_after_operation(time_before_move_mouse_and_smooth_scroll)  # 移动鼠标及窗口，前静置时间
    try:
        ActionChains(browse_driver).move_to_element(element_to_be_click).perform()
        browse_driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});",
                                     element_to_be_click)
        wait_after_operation(time_after_move_mouse_and_smooth_scroll)  # 移动鼠标及窗口，后等待时间
        element_to_be_click.click()
        print(f'移动鼠标位置至控件：{name_of_element}，点击成功！')
        return True
    except Exception as e:
        message = f'使用移动鼠标位置至控件：{name_of_element}，进行点击失败！'
        to_qw_error(message, exception_info=e, is_debug=True)
        return False

#将窗口关闭至只剩下一个窗口
def close_window_to_one(browse_driver, window_name):
    """
    将窗口关闭至只剩下一个窗口
    :param browse_driver:
    :param window_name: 回显给用户的提示信息中的 窗口名
    :return:
    """
    browse_driver.switch_to.default_content()
    current_window_numer = len(browse_driver.window_handles)
    print(f"准备关闭{window_name}，当前窗口数量：{current_window_numer}")
    try:
        browse_driver.close()
        is_hyper_link_closed = judge_window_closed(browse_driver, current_window_numer - 1,
                                                            window_name=window_name)
    except Exception:
        is_hyper_link_closed = False
    wait_after_operation(cfg.time_after_close_window)

    browse_driver.switch_to.window(browse_driver.window_handles[-1])
    if not is_hyper_link_closed:
        print(f"程序关闭{window_name}失败，请手动关闭窗口至1个后，程序继续执行！")
        while len(browse_driver.window_handles) > 1:
            wait_after_operation(cfg.time_wait_user_operation_between_poll)
            browse_driver.switch_to.window(browse_driver.window_handles[-1])
        is_hyper_link_closed = True

    browse_driver.switch_to.window(browse_driver.window_handles[-1])
    if is_hyper_link_closed and len(browse_driver.window_handles) > 1:
        index = len(browse_driver.window_handles)
        while index > 1:
            index -= 1
            try:
                browse_driver.close()
            except Exception:
                print("关闭{window_name}出现异常,需要手动关闭!")
                break
            else:
                browse_driver.switch_to.window(browse_driver.window_handles[-1])

    browse_driver.switch_to.window(browse_driver.window_handles[-1])
    if len(browse_driver.window_handles) > 1:
        print(f"程序出现异常，请手动关闭窗口至1个后，程序继续执行！")
        while len(browse_driver.window_handles) > 1:
            wait_after_operation(cfg.time_wait_user_operation_between_poll)
            browse_driver.switch_to.window(browse_driver.window_handles[-1])


#将窗口关闭至指定的数量
def close_window_to_specify_number(browse_driver, window_name, target_number):
    """
    将窗口关闭至指定的数量
    :param browse_driver:
    :param window_name: 回显给用户的提示信息中的 窗口名
    :param target_number:需要窗口达到的目标数：
    :return:
    """
    browse_driver.switch_to.default_content()
    current_window_numer = len(browse_driver.window_handles)
    if current_window_numer <= target_number:
        print(f"当前窗口数量：{current_window_numer}，小于目标数量：{target_number}。直接返回！")
        return
    print(f"准备关闭{window_name}，当前窗口数量：{current_window_numer}")

    is_window_closed = True
    while current_window_numer > target_number and is_window_closed:
        try:
            browse_driver.close()
            is_window_closed = judge_window_closed(browse_driver, current_window_numer - 1,
                                                            window_name=window_name)
        except Exception:
            is_window_closed = False
            break
        browse_driver.switch_to.window(browse_driver.window_handles[-1])
        browse_driver.switch_to.default_content()
        current_window_numer = len(browse_driver.window_handles)
    wait_after_operation(cfg.time_after_close_window)

    browse_driver.switch_to.window(browse_driver.window_handles[-1])
    browse_driver.switch_to.default_content()
    if not is_window_closed:
        print(f"程序关闭{window_name}失败，请手动关闭窗口至{target_number}个后，程序继续执行！")
        while len(browse_driver.window_handles) > target_number:
            wait_after_operation(cfg.time_wait_user_operation_between_poll)
            browse_driver.switch_to.window(browse_driver.window_handles[-1])
        is_window_closed = True

    browse_driver.switch_to.window(browse_driver.window_handles[-1])
    if is_window_closed and len(browse_driver.window_handles) > target_number:
        index = len(browse_driver.window_handles)
        while index > target_number:
            index -= 1
            try:
                browse_driver.close()
            except Exception:
                print("关闭{window_name}出现异常,需要手动关闭!")
                break
            else:
                browse_driver.switch_to.window(browse_driver.window_handles[-1])

    browse_driver.switch_to.window(browse_driver.window_handles[-1])
    if len(browse_driver.window_handles) > target_number:
        print(f"程序出现异常，请手动关闭窗口至{target_number}个后，程序继续执行！")
        while len(browse_driver.window_handles) > target_number:
            wait_after_operation(cfg.time_wait_user_operation_between_poll)
            browse_driver.switch_to.window(browse_driver.window_handles[-1])

#browse_driver当前激活页进行分析，使用的方法。对该元素的内容进行提取
#调用函数前，需确认element_id 已完成加载。
def get_content_by_driver_wait(web_browse_driver, By_method, By_param, max_wait_seconds = 50.0):
    """
    首先等待 element_id 对应的 element加载完成， 然后进行提取内容。 若加载超时，返回”人工查询内容“
    :param web_browse_driver: webdriver 。
    :param By_method: 使用的By参数
    :param By_param：与 By_method 对应的参数
    :param max_wait_seconds: 最大等待时间
    :return:element_id 的内容， full_content, shorten_content, 若element_id 内部包含表格，则shorten_content为表格内容
    """
    #判断 element_id 是否加载完成
    start_time = time.time()
    try:
        content_element = WebDriverWait(web_browse_driver, max_wait_seconds).until(EC.visibility_of_element_located((By_method, By_param)))
    except TimeoutException:
        # 加载页面失败：
        print(f'加载弹出页面失败！ 对应内容需人工查询')
        full_content = "人工查询内容"
        return full_content
    print(f"窗口内容可见用时：{time.time() - start_time}")

    # 获取页面内正文内容的标签
    full_content = "无正文"
    # 使用只有一次的循环语句，实现可在条件不满足时随时跳过后续语句执行
    for _index in [1]:
        #检查是否存在正文内容，并提取内容
        try:
            full_content = content_element.text
        except Exception as e:
            # 找不到内容部分，退出循环
            print(f"{By_method}：{By_param} 的内容无法读取!\n{e}")
            break
        # 查看元素中是否有table元素
        if full_content is None or len(full_content) < 1:
            full_content = f'页面内容无法智能提取'
            break
    return full_content


########################################################################################################################
#判断交互操作结果

#判断窗口是否被打开,通常用于判断点击超链接后新窗口是否打开
def judge_window_opened(browse_driver, success_windows_number, window_name="new window", retry_times=10, wait_seconds=1.0)->bool:
    """
    判断窗口是否成功生成
    :param browse_driver:
    :param success_windows_number: 如果成功，窗口数量。
    :param window_name:
    :param retry_times: 重试次数
    :param wait_seconds:每次等待时间
    :return:
    """
    if browse_driver is None :
        print("browse_driver is None!")
        return False
    try_index = 0
    is_open_succeed = False
    while try_index < retry_times and not is_open_succeed:
        try_index += 1
        if len(browse_driver.window_handles) >= success_windows_number:
            is_open_succeed = True
        else:
            try:
                WebDriverWait(browse_driver, wait_seconds ).until(EC.number_of_windows_to_be(success_windows_number))
                is_open_succeed = True
            except Exception as e:
                print(f"等待打开窗口：{window_name} 第{try_index}次等待超时!")
    if is_open_succeed:
        print(f"窗口：{window_name} 成功打开!")
    else:
        print(f"打开窗口：{window_name} 失败!")
    return is_open_succeed

#判断窗口是否被关闭,通常用于判断点击关闭窗口操作
def judge_window_closed(browse_driver, succeed_windows_number, window_name="new window", retry_times=10, wait_seconds=1.0)->bool:
    """
    判断窗口是否成功关闭
    :param browse_driver:
    :param succeed_windows_number: 如果成功关闭窗口后的窗口数量。
    :param window_name: 带关闭的窗口
    :param retry_times: 重试等待次数
    :param wait_seconds:
    :return:
    """
    if browse_driver is None :
        print("browse_driver is None!")
        return False
    try_index = 0
    is_close_succeed = False
    while try_index < retry_times and not is_close_succeed:
        try_index += 1
        if len(browse_driver.window_handles) <= succeed_windows_number:
            is_close_succeed = True
        else:
            try:
                WebDriverWait(browse_driver, wait_seconds).until(EC.number_of_windows_to_be(succeed_windows_number))
                is_close_succeed = True
            except Exception  as  e:
                print(f"等待关闭窗口：{window_name} 第{try_index}次等待超时!")
    if is_close_succeed:
        print(f"窗口：{window_name} 成功关闭!")
    else:
        print(f"关闭窗口：{window_name} 失败!")
    return is_close_succeed

#等待页面加载完成
#maxseconds=0, 表示持续等待直至页面成功加载。0 无时间长度限制.
def wait_document_ready(web_browse_driver, max_seconds=20, force_wait_time=0):
    """
    等待浏览器完成数据加载
    :param web_browse_driver:
    :param max_seconds: 等待页面加载完成的最长等待时间
    :param force_wait_time: 强制至少等待的时间长度
    :return:
    """
    start_time= time.time()
    while True:
        if "complete" == web_browse_driver.execute_script("return document.readyState"):
            break
        else:
            sleep(1) #等待1秒
            max_seconds-=1
            if max_seconds == 0:
                break
    if force_wait_time>0:
        delta_time = force_wait_time -(time.time() - start_time)
        if delta_time > 0.001:
            sleep(delta_time)
    return


########################################################################################################################
#人工智能
#使用用户定义的提示词进行内容分析
def AI_analysis_content_with_prompt(user_content, user_define_sys_prompt)->json:
    """
    按照用户定义的提示词，对content内容进行分析
    :param user_content: 待分析的内容
    :param user_define_sys_prompt: 用户定义的系统提示词
    :return:  json
    """
    client = OpenAI(
        base_url="https://api.deepseek.com/",
        api_key="sk-c549e0e1692844a7aeac77dc5eb92b5c"
    )

    completion = client.chat.completions.create(
        model="deepseek-chat",
        messages=[
            {
                "role": "system",
                "content": user_define_sys_prompt},
            {
                "role": "user",
                "content": user_content
            }
        ],
        response_format={
            'type': 'json_object'
        }
    )
    return json.loads(completion.choices[0].message.content)


########################################################################################################################
#输出

#报警函数,debug 用于是否打印报错信息
def to_qw_error(message,exception_info=None, is_debug= False):
    """
    :param message: 需要打印的信息
    :param exception_info: 异常信息
    :param is_debug: 默认False, 可以去强制打开
    :return:
    """
    print(message)
    if is_debug and (exception_info is not None):
        print(exception_info)
    return


########################################################################################################################
#文件操作

#获取excel work_sheet 中的第一列不为空的一行作为列名行。 的list
def get_excel_sheet_column_names(work_sheet:Worksheet) ->list:
    column_name_list = []
    for index in range(1, work_sheet.max_column+1):
        column_name_list.append(f'{work_sheet.cell(row=1,column= index).value}')  #强制类型转换，防止出现非字符串类型，数字类型等报错
    return column_name_list


#从excel 文件中读入待查公司名称列表，及名称的查询的附加条件
def get_company_list(company_file) -> numpy.ndarray:
    """
    从指定Excel文件中读取单位列表
    必须具备 排除查询、公司名称、别名、排除词组 列
    :param company_file:  存储单位列表的 excel 文件
    :return:  返回 单位 二维数组，3列 公式名称， 别名，排除词组
    """
    company_date_frame = pd.read_excel(company_file)
    company_date_frame['排除查询'] =  company_date_frame['排除查询'].astype(str)
    company_date_frame = company_date_frame.loc[company_date_frame['排除查询'] != '是', ["公司名称","别名","排除词组","销售部门","客户代表"]]
    return company_date_frame.to_numpy()


#根据标准排列standard_list, 将source_list,中的元素转换为在standard_list的位置序号。位置序号从1开始
# source_list的元素可以是数字或字符串，或两种形式混排。若出现不存在的序号，或者字符串，均返回false。成功为true
def transform_list_to_int_order_list(source_list:list, standard_list:list, int_order_list:list, is_check_int_in_source_list = True) ->bool:
    """
    将source_list中的元素，转换为在standard_list中对应远射位置的的列表。结果存储在int_order_list中。
    :param source_list:  待获取位置的字符序列。允许是数字字符串混排。
    :param standard_list: 标准位置排列字符串列表。 所有元素只能是字符串
    :param int_order_list: 返回排序的结果。 source_list中的元素若在 standard_list 不存在，对应位置设置为0 .
    :param is_check_int_in_source_list: True,检查source_list中的整数类型元素值域。[0, len(standard_list)]。
     source_list 中的元素，若在 standard_list 不存在对应位置设置为0 。 有效值域为[1, len(standard_list)] 。 0对应的变量在excel文件中将被舍弃。
    :return: 是否转换成功。
    """
    int_order_list.clear()
    #首先检查是否都是数字。
    is_all_int_in_source_list = all(isinstance(one_item,int) for one_item in source_list)
    #如果source_list中存在非数字的的元素，将其转换为standard_list的位置序号
    if not is_all_int_in_source_list:
        standard_list_dict={}
        for index in range(1, len(standard_list) + 1):
            standard_list_dict[standard_list[index - 1]] = index

        try:#防止source_list中出现 standard_list中不存在的字符。 或者其他非字符串类型，非整数类型等情况
            for one_item in source_list:
                if type(one_item) is int:
                    int_order_list.append(one_item)
                else:
                    try:
                        int_order_list.append(standard_list_dict[one_item])
                    except KeyError:   #对于不存在的键值，返回0
                        int_order_list.append(0)
        except Exception:
            print(f"transform_list_to_order_map中的source_list包含不正确的数据！\nsource_list is:\n{source_list} ")
            return False
    else: #若source_list全为整数，逐项复制 source_list中的值
        for one_item in source_list:
            int_order_list.append(one_item)

    # 如果不需检查element in int_order_list 的值域
    if not is_check_int_in_source_list:
        return True
    # 如果不需检查element in int_order_list 的值域
    # 检查是否存在超过standard_list长度的整数
    max_value = len(standard_list)
    for int_value in int_order_list:
        if int_value>max_value or int_value<0:
            return False
    return True

#按照指定的日期范围规则，读取 excel_filename中required_column_list列的数据。
def get_dataframe_from_excel_file_according_date(excel_filename, required_column_list, date_column_index, start_date, end_date,  sheet_name =None,sheet_index=0, start_row_number = None, end_row_number=None):
    """
    返回再日期范围内的数据， 没有符合要求的数据时，返回None
    :param excel_filename:
    :param required_column_list:  该列表中，应该包括日期列。
    :param date_column_index:  日期列 在 required_column_list 所在的序号。起始序列号0
    :param start_date: date 数据类型
    :param end_date:  date 数据类型
    :param sheet_name:excel 文件中，数据所在的sheet名,
    :param sheet_index:excel 文件中，数据所在的sheet序列号,
    :param start_row_number:数据起始行
    :param end_row_number: 数据结束行
    :return:返回满足日期要求的 dataframe 。若无满足要求的数据，返回None
    """

    excel_work_book = load_workbook(excel_filename)
    if sheet_name is None:
        excel_work_sheet = excel_work_book.worksheets[sheet_index]
    else:
        excel_work_sheet = excel_work_book[sheet_name]
    #转换target_column_list为纯数字序列。并检验数字范围的合规性

    int_column_list=[]
    # 获取标题行
    sheet_column_name_list = [cell.value for cell in excel_work_sheet[1]]
    if not transform_list_to_int_order_list(required_column_list, sheet_column_name_list, int_column_list):
        #转换中存在错误， 报错并退出
        print(f'get_content_from_excel_file_according_date 参数required_column_list 存在不正确的值！')
        return None

    if start_row_number is None:
        start_row_number = 2
    if end_row_number is None:
        end_row_number = excel_work_sheet.max_row

    if start_row_number > end_row_number:
        return None

    date_col_name = required_column_list[date_column_index]
    try:
        date_col_idx = sheet_column_name_list.index(date_col_name) + 1   # 转换为1-based索引
    except ValueError as e:
        raise ValueError(f"日期列： '{e.args[0]}' 不存在于Excel文件中") from e

    result_counter = 0
    result_data = []
    for row_idx in range(start_row_number, end_row_number+1):
        date_cell_value = excel_work_sheet.cell(row=row_idx, column=date_col_idx).value
        date_cell_value = f'{date_cell_value}'
        date_value = string_to_datetime(date_cell_value).date()
        if date_value > end_date or date_value < start_date :
            continue
        result_counter += 1
        temp_row = {}
        for index in range(0,len(required_column_list)):
            temp_row[required_column_list[index]] = f"{excel_work_sheet.cell(row= row_idx, column= int_column_list[index]).value}"
        #把日期列对应的数据修改为日期
        temp_row[date_col_name] = date_value
        result_data.append(temp_row)

    excel_work_book.close()
    df = None
    if result_counter >0 :
        df = pd.DataFrame(result_data)
    return df



# 初始化查询记录文件。无论是新生成查询记录文件，还是使用历史查询文件， 对于文件不存在的情况，都要新建文件
def initialize_record_file(excel_filename, is_new_query_file, column_list,sheet_name=None) -> bool:
    """
    初始化查询记录文件。无论是新生成查询记录文件，还是使用历史查询文件， 对于文件不存在的情况，都要新建文件
    :param excel_filename:  存储查询记录的文件名
    :param sheet_name:  存储查询记录的 sheet_name
    :param is_new_query_file:  是否是新生成的查询文件 。对非统计状态有效！
    :param column_list:  存储文件的第一行的列名。 仅对于新生成文件时有效
    :return:  返回文件初始化成功。
    """
    #由于statistic  情况下， 一直使用同一个文件。 非statistic 情况下， 文件名与重复概率较低，故注销下面代码。
    """
    if not is_new_query_file:  # 如是继续上一次的检查，故历史查询记录文件理论上应该已经存在，需检查该文件是否存在
        if os.path.exists(excel_filename):
            return True
    """
    #检查查询文件是否存在，不存在则生成文件
    if os.path.exists(excel_filename):
        if is_new_query_file:
            print(f"提示：文件:{excel_filename} 已经存在！ 统计状态可忽略！")
        return True
    #文件不存在，需要创建新文件
    excel_workbook = Workbook()
    if sheet_name is None:
        work_sheet = excel_workbook.active
    else:
        if sheet_name in excel_workbook.sheetnames:
            work_sheet = excel_workbook[sheet_name]
        else:
            excel_workbook.create_sheet(title=sheet_name, index=0)
            work_sheet = excel_workbook.worksheets[0]
    work_sheet.append(column_list)
    excel_workbook.save(excel_filename)
    if not is_new_query_file:
        print(f"对于非新生成查询，文件应该已存在。本次未查询到文件：{excel_filename}。已重新生成。")
    return True




#从excel文件中读取未分析的数据，取出 ["标题", "业务相关性", "公告摘要", "公告正文"]4列。
def get_unanalyzed_records_from_excel_file( excel_filename, required_column_list, relevance_index, sheet_name =None, sheet_index=0, max_return_row_amount = 10, is_reverse_order=True, start_row_number = None):
    """
    按照required_column_list中指定的列明很取数数
    :param excel_filename: 需要被读写的文件。
    :param required_column_list:["标题", "业务相关性", "公告摘要", "公告正文"]
    :param relevance_index: 业务相关性列在 required_column_list 中的索引。 从0 开始
    :param sheet_name: 如果指定sheet_name , 使用sheet_name 。优先级高于 sheet_index
    :param sheet_index: 默认访问第一个sheet
    :param max_return_row_amount: 最大返回的行数，默认为10
    :param is_reverse_order: 是否逆序，True从最后一行开始匹配。False，从第一行开始。默认为逆序
    :param start_row_number 检索的起点。
    :return: result_rows 2维list, 第一个数据为行号。 行号， [对照 required_column_list 的 数据]； [[1,[v1,v2,v3...]]]。若 target_column_list设置错误，返回值为 None！！！
    """
    #不需要处理的行

    excel_work_book = load_workbook(excel_filename)
    if sheet_name is None:
        excel_work_sheet = excel_work_book.worksheets[sheet_index]
    else:
        excel_work_sheet = excel_work_book[sheet_name]
    #转换target_column_list为纯数字序列。并检验数字范围的合规性
    int_column_list = []
    # 获取标题行
    sheet_column_name_list = [cell.value for cell in excel_work_sheet[1]]
    if not transform_list_to_int_order_list(required_column_list, sheet_column_name_list, int_column_list):
        #转换中存在错误， 报错并推出
        print(f'get_unanalyzed_records_from_excel_file 参数required_column_list 存在不正确的值！')
        return None

    if start_row_number is None:
        #对于空值情况，给start_row_number 赋值
        if is_reverse_order:
            start_row_number = excel_work_sheet.max_row
        else:
            start_row_number = 2
    else:
         # 对于非空值情况，检查start_row_number 值域的合法性。不合法直接退出。
        if start_row_number > excel_work_sheet.max_row or start_row_number < 2:
            return []
    #确定行序范围
    if is_reverse_order:
        row_range = range(start_row_number,1,-1)
    else:
        row_range = range(start_row_number,excel_work_sheet.max_row+1)
    #按照行序开始检索文件
    result_rows = []
    result_counter = 0
    for row_index in row_range:
        #为防止不是字符串的情况，如 数字、空单元格等，强制转换为 string 进行比较。但返回值不做调整
        relevance_cell_value = excel_work_sheet.cell(row= row_index, column=int_column_list[relevance_index]).value
        if relevance_cell_value is not None  and  len(relevance_cell_value.strip()) > 0: #
            continue  #已经分析过的内容直接跳过
         #未分析的内容，需要提取
        temp_row = []
        for index in range(0,len(int_column_list)):
            temp_row.append( f'{excel_work_sheet.cell(row= row_index, column=int_column_list[index]).value}')
        temp_row[relevance_index] = ""  #给相关性赋空值。
        result_rows.append([row_index,temp_row])
        result_counter += 1
        if result_counter >= max_return_row_amount :
            break
    #关闭文件，并返回结果
    excel_work_book.close()
    return result_rows


def do_update_static_info(query_string, analysis_data_column_list, function_point, analysis_result_column, in_reverse_direction, max_result_number, max_quota):
    """
    对数据进行更新，
    :param query_string:  具体更新的数据类型， 包括：business_opportunity等，相关定义见： config.global_query_types
    :param analysis_data_column_list:  更新分析需要获取的行， 目前仅仅支持：["标题","业务相关性", "公告摘要", "公告正文"]
    :param function_point:  对于本类型数据进行分析更新函数。 返回值必须为 Bool  , 调用格式为：funciton_point(one_param)
    :param analysis_result_column： 分析结果会写的列
    :param in_reverse_direction:  在文件中更新的顺序
    :param max_result_number: 每次检索出的需要更新的数据条数
    :param max_quota:   最大累计可更新的条目数
    :return:
    """

    query_type =query_string
    column_in_excel = analysis_data_column_list
    column_in_excel = ["标题", "业务相关性", "公告摘要", "公告正文"]
    file_name_to_update_info, datetime_begin, datetime_end, query_company_list, is_last_finished = get_statistic_history_information(
        cfg.filename_history_query_log, query_type, release_date_type= "date")

    if is_last_finished:
        print(f'本次更新文件: {file_name_to_update_info}, 本文件信息采集至：{datetime_begin}')
    else:
        print(f'本次更新文件: {file_name_to_update_info}, 尚未完成：{datetime_begin}---{datetime_end} 的信息采集！')
    #开始数据更新
    start_row_number = None #初始化条件
    used_quota = 0
    # 存储刷新成功的数据结果
    updated_records_list = []
    while used_quota < max_quota :
        match_records_list = get_unanalyzed_records_from_excel_file( file_name_to_update_info, column_in_excel,1,
                                                                             max_return_row_amount= max_result_number, is_reverse_order= in_reverse_direction, start_row_number= start_row_number)
        match_records_list_length = len(match_records_list)
        if match_records_list_length == 0:
            print(f"在{query_type}中，未检索到需要更新的数据")
            break #退出while 循环
        else:
            #为下一轮循环配置起点。
            start_row_number = match_records_list[match_records_list_length - 1][0]
            if in_reverse_direction:
                start_row_number -= 1
            else:
                start_row_number +=1


        # 开始处理返回的每一个行。
        #首先定义初始变量
        from base_function_zfcg import Struct_purchase_intention
        from base_function_jianyu import Struct_business_opportunity
        if query_type.startswith("business_opportunity"):
            temp_struct_record=Struct_business_opportunity()
        elif query_type.startswith("purchase_intention"):
            temp_struct_record = Struct_purchase_intention()
        else:
            print("不支持的数据类型，程序退出！")
            exit(cfg.ErrorCode_not_support_query_type)
        for one_row in match_records_list:
            temp_struct_record.re_init()
            temp_struct_record.title = one_row[1][0]
            temp_struct_record.relevance = ""
            temp_struct_record.summary_info = one_row[1][2]
            temp_struct_record.complete_info = one_row[1][3]
            is_current_row_update_success = function_point(temp_struct_record)
            if is_current_row_update_success :
                print(f"第{one_row[0]}行完成分析！")
            else:
                print(f"第{one_row[0]}行分析失败！")
            #判断本行数据分析是否成功
            if is_current_row_update_success: #本行重刷数据成功，将更新后的如数压入updated_records_list
                one_success_record = [one_row[0], temp_struct_record.push_data_into_list_according_column_list(analysis_result_column)]
                updated_records_list.append(one_success_record)
                used_quota += 1
                if used_quota >= max_quota :
                    break #到达更新限额,停止for 循环
        #退出for 循环，若有数据分析，需要在源文件中更新
        updated_success_number = len(updated_records_list)
        if updated_success_number > 0: #本轮有成功更新数据，
            is_update_file_success = insert_record_list_into_file(file_name_to_update_info,updated_records_list , analysis_result_column)
            updated_records_list.clear() #插入后清空数据
            if not is_update_file_success:
                print(f'向文件：{file_name_to_update_info} 更新内容失败！，请检查是否文件未关闭！累计已更新{used_quota}条记录')
                exit(cfg.ErrorCode_read_write_file)
            else:
                print(f'向文件：{file_name_to_update_info} 更新 {updated_success_number} 条记录成功！累计已更新{used_quota}条记录')

        if  used_quota>= max_quota:
            print(f'今日更新额度已用完！分配配额 {max_quota}, 实际使用配额{used_quota}')
            break  #停止while循环

#把记录增加到excel_filename 文件的末尾
def append_record_list_into_file(excel_filename,  data_list,column_list = None, sheet_name=None, sheet_index = 0, is_check_column_list_int_value_range= True) -> bool:
    """
    向excel_filename文件添加data_list中记录的信息
    :param excel_filename:
    :param data_list:二维list
    :param column_list: 把数据按照对应的 column_list ,添加到文件末尾行。column-list  可以是数字或文字，或混合模式，均可
    :param sheet_name： sheet 名字，如果名字在sheet中不存在，返回False
    :param sheet_index, 如果未提供 sheet_name ,使用 sheet_index 如果sheet_index ,大于sheet数-1 ， 返货回 False
    :param is_check_column_list_int_value_range: 默认值：True;若 column_list 中存在整数，需要检查其值是否在当前表的列的宽度范围内。[1, work_sheet.max_column]。
    :return: 返回添加记录操作是否成功
    """
    try:
        excel_workbook =load_workbook(excel_filename)
        if sheet_name is None:
            if sheet_index > len(excel_workbook.sheetnames)-1: #sheet_index 超越该表最大允许 index
                excel_workbook.close()
                return False
            work_sheet = excel_workbook.worksheets[sheet_index]
        else:
            if sheet_name  in excel_workbook.sheetnames:
                work_sheet = excel_workbook[sheet_name]
            else : #sheet_name 不存在
                excel_workbook.close()
                return False
        if column_list is None :  #不存在column_list 限制
            for data_row in data_list:
                work_sheet.append(data_row)
        else: #存在column_list 限制,按照存在column_list指定“列表”进行插入
            int_column_list=[]
            column_name_list_in_excel_sheet = get_excel_sheet_column_names(work_sheet)
            if len(column_name_list_in_excel_sheet) <1 :
                print(f"append_record_list_into_file sheet表名不存在，或者 指定的的 column_list不存在！\n{column_list}")
                exit(cfg.ErrorCode_wrong_sheet_column_list)
            if not transform_list_to_int_order_list(column_list,column_name_list_in_excel_sheet,int_order_list=int_column_list, is_check_int_in_source_list= is_check_column_list_int_value_range):
                #存在超出最大列数的列编号，退出
                print(f"append_record_list_into_file column list 中存在错误的列序号！\n{column_list}")
                exit(cfg.ErrorCode_sheet_column_order_overflow)
            current_row = work_sheet.max_row
            for one_row_data in data_list:
                current_row += 1
                for column_index in range(0,len(int_column_list)):
                    if int_column_list[column_index]<1:  #列序号为0的元素不屑于表格择偶那个
                        continue
                    work_sheet.cell(row=current_row,column=int_column_list[column_index]).value=one_row_data[column_index]
        excel_workbook.save(excel_filename)
        return True
    except Exception as e:
        to_qw_error(message=f"向文件：{excel_filename} 添加记录出错！请检查该文件是否被其他应用程序打开。或者column_list存在错误", exception_info=e)
        exit(cfg.ErrorCode_read_write_file)
        #return False


#把记录插入到excel_filename 文件中指定的位置。 行位置，由data_list提供， 列位置由 column 提供。
def insert_record_list_into_file(excel_filename,  data_list,column_list , sheet_name=None, sheet_index = 0, is_check_column_list_int_value_range = True) -> bool:
    """
    向excel_filename文件添加data_list中记录的信息
    :param excel_filename:
    :param data_list:# 格式： [[行号1，[dta11,data12...]],[行号2，[dta21,data22...]...] ]
    :param column_list: 把数据按照对应的 column_list ,添加到文data_list指定行号  可以是数字或文字，或混合模式均可
    :param sheet_name： sheet 名字，如果名字在sheet中不存在，返回False
    :param sheet_index, 如果未提供 sheet_name ,使用 sheet_index 如果sheet_index ,大于sheet数-1 ， 返货回 False
    :param is_check_column_list_int_value_range: 默认值：True;若 column_list 中存在整数，需要检查其值是否在当前表的列的宽度范围内。[1, work_sheet.max_column]
    :return: 返回添加记录操作是否成功
    """
    try:
        excel_workbook =load_workbook(excel_filename)
        if sheet_name is None:
            if sheet_index > len(excel_workbook.sheetnames)-1: #sheet_index 超越该表最大允许 index
                excel_workbook.close()
                return False
            work_sheet = excel_workbook.worksheets[sheet_index]
        else:
            if sheet_name  in excel_workbook.sheetnames:
                work_sheet = excel_workbook[sheet_name]
            else : #sheet_name 不存在
                excel_workbook.close()
                return False

        int_column_list=[]
        column_name_list_in_excel_sheet = get_excel_sheet_column_names(work_sheet)
        if len(column_name_list_in_excel_sheet) <1 :
            print(f"insert_record_list_into_file sheet表名不存在，或者 指定的的 column_list不存在！\n{column_list}")
            exit(cfg.ErrorCode_wrong_sheet_column_list)
        if not transform_list_to_int_order_list(column_list,column_name_list_in_excel_sheet,int_order_list=int_column_list, is_check_int_in_source_list= is_check_column_list_int_value_range):
            #存在超出最大列数的列编号，退出
            print(f"insert_record_list_into_file column list 中存在错误的列序号！\n{column_list}")
            exit(cfg.ErrorCode_sheet_column_order_overflow)
        for one_row_data in data_list:
            current_row = one_row_data[0]
            for locator_index in range(0, len(int_column_list)):
                if int_column_list[locator_index] < 1:  # 列序号为0的元素不屑于表格择偶那个
                    continue
                work_sheet.cell(row=current_row,column=int_column_list[locator_index]).value=one_row_data[1][locator_index]
        excel_workbook.save(excel_filename)
        return True
    except Exception as e:
        to_qw_error(message=f"向文件：{excel_filename} 插入记录出错！请检查该文件是否被其他应用程序打开。或者column_list存在错误", exception_info=e)
        exit(cfg.ErrorCode_read_write_file)
        #return False

#对指定位置的数据进行数据累加。
def update_statistic_data(excel_filename, data_list, column_name_list=None, row_number = 1, offset=1, sheet_name=None, sheet_index = 0):
    """
    使用 data_list 中的的数据 更新 colunn_name_list 对应的统计数据。
    :param excel_filename:
    :param data_list:
    :param column_name_list: 需要更新的数据名。 如果为None , 默认使用：["累计总记录","累计有效记录","累计相关记录"]
    :param row_number:  数据填写在第几行
    :param offset:  写入数据的位置，相对于column_name的偏移值
    :param sheet_name: excel_filename 中的 sheet 名
    :param sheet_index: excel_filename 中的 sheet index
    :return: True , 更新数据成功， FaLse ，更新失败
    """
    if column_name_list is None:
        column_list = ["累计总记录","累计有效记录","累计相关记录"]
    else :
        column_list = [e for e in column_name_list]
    try:
        excel_workbook =load_workbook(excel_filename)
        if sheet_name is None:
            if sheet_index > len(excel_workbook.sheetnames)-1: #sheet_index 超越该表最大允许 index
                excel_workbook.close()
                return False
            work_sheet = excel_workbook.worksheets[sheet_index]
        else:
            if sheet_name  in excel_workbook.sheetnames:
                work_sheet = excel_workbook[sheet_name]
            else : #sheet_name 不存在
                excel_workbook.close()
                return False

        int_column_list=[]
        column_name_list_in_excel_sheet = get_excel_sheet_column_names(work_sheet)
        if len(column_name_list_in_excel_sheet) <1 :
            print(f"update_statistic_data sheet表名不存在")
            exit(cfg.ErrorCode_wrong_sheet_column_list)
        transform_list_to_int_order_list(column_list,column_name_list_in_excel_sheet,int_order_list=int_column_list)
        #检查int_column_list 中是否存在0 ，若有， 表明错误的列名
        for index in range(0,len(int_column_list)):
            if int_column_list[index] == 0:
                print(f"update_statistic_data column_name_list 中存在错误的列名！\n{column_list}")
            else:
                int_column_list[index] += offset

        for locator_index in range(0, len(int_column_list)):
            if int_column_list[locator_index] < 1:  # 列序号为0的元素不屑于表格择偶那个
                continue
            work_sheet.cell(row=row_number,column=int_column_list[locator_index]).value += data_list[locator_index]
        excel_workbook.save(excel_filename)
        return True
    except Exception as e:
        to_qw_error(message=f"向文件：{excel_filename} 插入记录出错！请检查该文件是否被其他应用程序打开。或者data_list存在错误", exception_info=e)
        exit(cfg.ErrorCode_read_write_file)
        #return False



########################################################################################################################
#历史操作信息获取
#创建日志文件
def create_log_file(log_file_name):
    """
    :param log_file_name: 日志文件的名称
    :return: 是否创建日志文件成功
    """
    # 检查查询记录历史文件是否存在
    if os.path.exists(log_file_name):
        return True
    # log文件不存在，则创建文件
    begin_point = datetime.strptime("1999-01-01", "%Y-%m-%d")
    end_point = datetime.strptime("2000-01-01", "%Y-%m-%d")
    datetime_point = datetime.now()
    excel_workbook = Workbook()
    for index in range(0,len(global_query_types)):
        excel_workbook.create_sheet(title = global_query_types[index], index=index)
        active_sheet = excel_workbook.worksheets[index]
        if global_query_types[index].endswith("_statistic"):  #统计类型
            active_sheet.append(["公司名称", "总记录数","有效记录数","相关记录数", "时间", "时段起点", "时段终点","累计记录数",0,"累计有效记录",0,"累计相关记录",0])
            temp_file_name = cfg.directory_to_save_statistic_record + "/" + global_query_Chinese_name[index] + ".xlsx"
            active_sheet.append([temp_file_name, 0,-1,0, datetime_point, begin_point, end_point])
        else:  #临时查询类型
            active_sheet.append(["公司名称", "有效记录数", "时间"])
            temp_file_name = cfg.directory_to_save_query_record + "/" + global_query_Chinese_name[index] + ".xlsx"
            active_sheet.append([temp_file_name, -1, datetime_point])

        #statistic 为长期文件， 需要关注每次查询的时间起点和终点
        #非statistic 为临时文件，不需要关注每次的时间起点。
        # 有效记录数  -3 : 表示继续上次未完成查询， 公司名称列为 被继续使用的存储文件名
        # 有效记录数  -2 : 表示开始查询， 公司名称列为 存储文件名
        # 有效记录数  -1 : 表示结束查询，
    excel_workbook.save(log_file_name)


def re_initialize_logfile(excel_filename):
    begin_point = datetime.strptime("1999-01-01", "%Y-%m-%d")
    end_point = datetime.strptime("2000-01-01", "%Y-%m-%d")
    datetime_point = datetime.now()
    excel_workbook = load_workbook(excel_filename)
    sheet_names = excel_workbook.sheetnames
    for index in range(0,len(global_query_types)):
        if global_query_types[index] not in sheet_names:
            excel_workbook.create_sheet(title= global_query_types[index], index = index)
            active_sheet = excel_workbook.worksheets[index]
            if global_query_types[index].endswith("_statistic"):  # 统计类型
                active_sheet.append(["公司名称", "总记录数","有效记录数","相关记录数", "时间", "时段起点", "时段终点","累计记录数",0,"累计有效记录",0,"累计相关记录",0])
                temp_file_name = cfg.directory_to_save_statistic_record + "/" + global_query_Chinese_name[
                    index] + ".xlsx"
                active_sheet.append([temp_file_name, 0,-1,0, datetime_point, begin_point, end_point])
            else:  # 临时查询类型
                active_sheet.append(["公司名称", "有效记录数", "时间"])
                temp_file_name = cfg.directory_to_save_query_record + "/" + global_query_Chinese_name[index] + ".xlsx"
                active_sheet.append([temp_file_name, -1, datetime_point])
    excel_workbook.save(excel_filename)




#针对统计场景下，获取已经完成的公司的名称，检测前次查询完成情况，如未完成，返回上一次文件名称。如已完成，返回本次新建查询文件名。参数增加了获取时间范围。
def get_statistic_history_information(query_log_filename, query_type, release_date_type="date", additional_info =None) -> (str,datetime, datetime,list, bool):
    """
    根据信息类型，选择读取不同的sheet,文件不存在则创建log文件
    注意，时间起点用永远是 日期+0点0分0秒
    :param query_log_filename:  指定的查询日志文件名
    :param query_type: 查询的类型， business_opportunity_statistic 等， 见 global_query_types
    :param release_date_type: 被查询的信息发布时间类型，从 global_release_date_types 选择
    :param additional_info: 为了函数的可可扩展性， 增加的返回信息 , 使用字典方案。返回本轮已完成"总记录数","有效记录数","相关记录数"汇总数据
    :return: 保存查询记录的文件名称 str，查询时间范围起点，查询时间范围终点，已经查询过的公司列表 list，本次是否为新查询任务。 bool
    """
    if release_date_type not in global_release_date_types:
        print( print(f'get_statistic_history_information支持的信息发布时间类型：{global_release_date_types}！'))
    if not query_type.endswith("_statistic") :
        print(f'get_statistic_history_information仅支持统计类型查询！')
        exit(cfg.ErrorCode_not_support_query_type)
    if query_type not in global_query_types:
        print(f'本程序支持的查询类型为：{global_query_types}。不支持您指定的查询类型：{query_type}！\n程序将退出！')
        exit(cfg.ErrorCode_not_support_query_type)

    save_statistic_record_filename = ""  # 保存查询结果的文件名
    queried_companies = []  # 已经被查询过的公司
    # 本次是否为一次新查询，false 表示本次为完成前次未完成的查询
    is_an_new_query = True
    #检查查询记录历史文件是否存在
    is_query_log_exist = os.path.exists(query_log_filename)
    if not is_query_log_exist:
        # log文件不存在，则创建文件
        create_log_file(query_log_filename)
    #存在历史记录文件（不包含在本次函数新创建的log 文件，其历史记录长度为0 ，无需要索引的内容），
    # 并且要求继续前次未完成的查询，才需要查询获得上一次上次查询是否完成，以及上次已经查询过的公司列表
    #读取内容
    date_range_begin = datetime.strptime("1990-01-01", "%Y-%m-%d")
    date_range_end = datetime.strptime("2000-01-01", "%Y-%m-%d")
    delta_hours = timedelta(hours=cfg.forward_hours) #设置时间时间前移值
    try:
        history_query_information = pd.read_excel(query_log_filename, sheet_name=query_type)
    except ValueError:
        print(f"{query_log_filename}中不存在sheet：{query_type}")
        re_initialize_logfile(query_log_filename)
        print(f"已成功初始化文件：{query_log_filename}")
        history_query_information = pd.read_excel(query_log_filename, sheet_name=query_type)
    history_query_information_length = len(history_query_information)
    total_records = 0
    total_available_records = 0
    total_relevance_records = 0
    if  history_query_information_length>0:
        loc_index = history_query_information_length-1
        while loc_index > -1:
            current_flag = history_query_information.iloc[loc_index]["有效记录数"]
            if current_flag == -1 :
                save_statistic_record_filename = history_query_information.iloc[loc_index]["公司名称"]
                date_range_begin = history_query_information.iloc[loc_index]["时段终点"]
                date_range_end = datetime.now()-delta_hours
                break
            elif current_flag == -2:
                save_statistic_record_filename = history_query_information.iloc[loc_index]["公司名称"]
                date_range_begin = history_query_information.iloc[loc_index]["时段起点"]
                date_range_end = history_query_information.iloc[loc_index]["时段终点"]
                is_an_new_query = False
                break
            elif current_flag >= 0:
                queried_companies.append( history_query_information.iloc[loc_index]["公司名称"])
                total_records += history_query_information.iloc[loc_index]["总记录数"]
                total_available_records += history_query_information.iloc[loc_index]["有效记录数"]
                total_relevance_records += history_query_information.iloc[loc_index]["相关记录数"]
            else:
                pass
            loc_index = loc_index -1
    #根据是否找到得到历史查询文件名，确认本次是新查询
    if len(save_statistic_record_filename) ==0  :  #未查询到历史查询文件
        is_an_new_query = True
        #防止出现log 文件被错误修改后，未查询到历史记查询录文件名，却保留了已查询公式名称记录。
        queried_companies=[]
        #生成新查询文件名称
        for index in range(0,len(global_query_types)):
            if query_type==global_query_types[index]: #匹配到对应的查询类型
                save_statistic_record_filename = global_query_Chinese_name[index] + ".xlsx"
                break
        date_range_begin = datetime.strptime("2000-01-01", "%Y-%m-%d")
        date_range_end = datetime.now()- delta_hours

    #将文件名同一更新到配置文件中指定的目录，如果以前存在的文件因配置文件的目录修改，也需要做调整
    save_statistic_record_filename = save_statistic_record_filename[save_statistic_record_filename.rfind("/")+1:]
    save_statistic_record_filename = cfg.directory_to_save_statistic_record+"/"+save_statistic_record_filename

    # 时段起点都化成当日的零点
    if release_date_type == "date":
        date_range_begin = date_range_begin.replace(hour=0, minute=0, second=0, microsecond=0)
    #打印本次查询的时间范围
    print(f"本次更新的统计数据类型为：{query_type}")
    if is_an_new_query:
        print(f"本次新增数据的时间范围为：{date_range_begin}----{date_range_end.replace(second=0, microsecond=0)}")
    else:
        print(f"本次继续获取数据的时间范围为：{date_range_begin}----{date_range_end.replace(second=0, microsecond=0)}")

    #本轮查询的相关数据汇总
    if  additional_info is not None and isinstance(additional_info, dict):
        additional_info.clear()
        additional_info["总记录数"] = int(total_records)
        additional_info["有效记录数"] =int(total_available_records)
        additional_info["相关记录数"]= int(total_relevance_records)
    # 返回结果
    return save_statistic_record_filename,date_range_begin,date_range_end.replace(second=0, microsecond=0),queried_companies,is_an_new_query




